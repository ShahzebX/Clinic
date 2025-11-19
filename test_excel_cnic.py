"""Test Excel export with empty CNIC."""

from datetime import datetime
from pathlib import Path
from database.db_manager import PatientRecord
from utils.excel_manager import ExcelManager
from utils.helpers import get_clinic_data_folder

# Create test record with empty CNIC
record = PatientRecord(
    date=datetime(2025, 11, 18),
    opd_no="TEST-001",
    name="Test Patient",
    father_name="Test Father",
    age=30,
    gender="Male",
    cnic=None,  # Empty CNIC
    address="Test Address Line 1\nTest Address Line 2",
    temperature=98.6,
    bp="120/80",
    weight=70.0,
    diabetic=100.0,
    fees_type="Normal",
)

# Create Excel manager and export
excel_manager = ExcelManager(get_clinic_data_folder())
path = excel_manager.append_record(record)

print(f"Excel file saved to: {path}")
print("\nRecord data:")
print(f"  OPD No: {record.opd_no}")
print(f"  CNIC: {repr(record.cnic)}")
print(f"  Address: {repr(record.address)}")

# Open the Excel file to verify
from openpyxl import load_workbook
wb = load_workbook(path)
ws = wb.active

# Get the last row
last_row = ws.max_row
print(f"\nLast row in Excel (row {last_row}):")
for col_idx, header in enumerate(ExcelManager.HEADERS, start=1):
    cell_value = ws.cell(last_row, col_idx).value
    print(f"  {header}: {repr(cell_value)}")
