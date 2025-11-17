"""Excel export functionality for the Clinic app."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from database.db_manager import PatientRecord
from utils.helpers import (
	current_month_filename,
	ensure_directory,
	format_date,
	format_fees,
)


class ExcelManager:
	"""Create and maintain monthly Excel workbooks for patient records."""

	HEADERS = [
		"OPD No",
		"Visit Date",
		"Patient Name",
		"Father/Husband Name",
		"Age",
		"Gender",
		"CNIC",
		"Address",
		"Temperature (°F)",
		"Blood Pressure",
		"Weight (kg)",
		"Diabetic (mg/dl)",
		"Fees Type",
	]

	def __init__(self, data_dir: Path | str) -> None:
		self.data_dir = ensure_directory(data_dir)

	def _workbook_path(self, timestamp) -> Path:
		filename = current_month_filename("ClinicData", timestamp)
		return self.data_dir / filename

	def append_record(self, record: PatientRecord) -> Path:
		"""Append *record* to the appropriate monthly workbook."""

		path = self._workbook_path(record.date)
		workbook, worksheet = self._get_or_create_workbook(path, record)

		worksheet.append(
			[
				record.opd_no,
				format_date(record.date),
				record.name,
				record.father_name,
				record.age,
				record.gender,
				record.cnic or "",
				record.address.replace("\n", ", ") if record.address else "",
				record.temperature,
				record.bp,
				record.weight if record.weight is not None else "",
				record.diabetic if record.diabetic is not None else "",
				format_fees(record.fees_type),
			]
		)
		self._autosize_columns(worksheet)
		workbook.save(path)
		return path

	def append_records(self, records: Iterable[PatientRecord]) -> None:
		"""Append multiple *records* efficiently."""

		grouped: dict[Path, list[list[object]]] = {}
		workbooks: dict[Path, tuple[Workbook, Worksheet]] = {}

		for record in records:
			path = self._workbook_path(record.date)
			if path not in workbooks:
				workbooks[path] = self._get_or_create_workbook(path, record)
				grouped[path] = []
			grouped[path].append(
				[
					record.opd_no,
					format_date(record.date),
					record.name,
					record.father_name,
					record.age,
					record.gender,
					record.cnic or "",
					record.address.replace("\n", ", ") if record.address else "",
					record.temperature,
					record.bp,
					record.weight if record.weight is not None else "",
					record.diabetic if record.diabetic is not None else "",
					format_fees(record.fees_type),
				]
			)

		for path, rows in grouped.items():
			workbook, worksheet = workbooks[path]
			for row in rows:
				worksheet.append(row)
			self._autosize_columns(worksheet)
			workbook.save(path)

	def _get_or_create_workbook(
		self, path: Path, record: PatientRecord
	) -> tuple[Workbook, Worksheet]:
		if path.exists():
			workbook = load_workbook(path)
			worksheet = workbook.active
		else:
			workbook = Workbook()
			worksheet = workbook.active
			worksheet.title = record.date.strftime("%B %Y")
			worksheet.append(self.HEADERS)
		return workbook, worksheet

	def _autosize_columns(self, worksheet: Worksheet) -> None:
		for idx, column_cells in enumerate(worksheet.columns, start=1):
			try:
				column = get_column_letter(idx)
			except ValueError:
				continue
			max_length = 0
			for cell in column_cells:
				value = "" if cell.value is None else str(cell.value)
				max_length = max(max_length, len(value))
			worksheet.column_dimensions[column].width = min(40, max_length + 2)


__all__ = ["ExcelManager"]
